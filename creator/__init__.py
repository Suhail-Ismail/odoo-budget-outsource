#!/usr/bin/env python

# IMPORT INITIALIZING DJANGO ORM
# IMPORT OPENPYXL WITH INSERT ROW
# excel password tbpc19
# ----------------------------------------------------------------------------------------------------
import os
import base64
import logging

_logger = logging.getLogger(__name__)

from openpyxl.styles import Protection

from .utils import *
from .utils import load_workbook, get_distinct_selection
import pandas as pd
import shutil

from odoo import fields


# MAIN COMPONENTS STARTS HERE
# ----------------------------------------------------------------------------------------------------
def odoo_to_pandas_list(orm_query=None, columns=list()):
    columns = columns if columns else orm_query.fields_get_keys()
    data = []
    for row in orm_query:
        row_data = {}
        for column in columns:
            row_data[column] = row.mapped(column)[0]
        data.append(row_data)
    return data


def get_required_hour(period_start=None, period_end=None, ramadan_start=None, ramadan_end=None):
    total_days = (period_end - period_start).days + 1
    ramadan_days = 0

    if ramadan_start and ramadan_end:
        if period_start <= ramadan_start <= period_end or period_start <= ramadan_end <= period_end:
            start = ramadan_start if period_start <= ramadan_start <= period_end else period_start
            end = ramadan_end if period_start <= ramadan_end <= period_end else period_end
            ramadan_days = (end - start).days + 1

    normal_days = total_days - ramadan_days

    # +1 is a corrections factor to count all days within a period
    return normal_days * 48.0 / 7.0 + ramadan_days * 36.0 / 7.0


def get_individual_required_hour(date_of_join=None, period_start=None, period_end=None,
                                 ramadan_start=None, ramadan_end=None):
    # date of date is latest
    if date_of_join > period_start:
        return get_required_hour(date_of_join, period_end, ramadan_start, ramadan_end)
    return get_required_hour(period_start, period_end, ramadan_start, ramadan_end)


def get_output_file(*args):
    return ' '.join(args) + '.xlsx'


class Creator(object):
    def __init__(self, env, form_name, xlsx_pass):
        self.env = env
        self.required_days = None
        self.period_start = None
        self.period_end = None
        self.ramadan_start = None
        self.ramadan_end = None
        self.total_required_hours = None
        self.period_string = None
        self.required_hour = None
        self.po_num = None
        self.context = None
        self.output_filename = None
        self.output_file_path = None
        self.division = None
        self.qs_po = None
        self.qs_resource = None
        self.qs_invoice = None
        self.qs_line_details = None
        self.qs_unit_price = None

        self.xlsx_pass = xlsx_pass

        self.zip_file = None
        self.base_dir = os.path.dirname(os.path.realpath(__file__))
        self.form = os.path.join(self.base_dir, 'forms_xlsx', form_name)

        self.temp_dir = tempfile.mkdtemp()
        self.output_dir = self.temp_dir

    def setup(self):
        self.period_string = '{0:%d-%b-%Y} to {1:%d-%b-%Y}'.format(self.period_start,
                                                                   self.period_end)
        # PERIOD MUST BE IN dd/mm/yyyy format()
        self.required_hour = self.total_required_hours if self.total_required_hours > 0 else \
            get_required_hour(self.period_start, self.period_end,
                              self.ramadan_start, self.ramadan_end)

        # Days
        self.required_days = (self.period_end - self.period_start).days + 1

    def make_zip(self):
        shutil.make_archive(base_name=self.output_dir,
                            format='tar',
                            root_dir=self.output_dir)
        self.zip_file = self.output_dir + '.tar'
        return self.zip_file

    @staticmethod
    def attach(env, res_id, res_model, file_path):
        # Attach generated document to filestore
        ir_attach = env['ir.attachment']
        full_path = os.path.join(file_path)

        with open(full_path, 'rb') as fp:
            data = base64.b64encode(fp.read())
        filename = os.path.split(full_path)[1]
        values = dict(
            name=filename,
            datas_fname=filename,
            res_id=res_id,
            res_model=res_model,
            type='binary',
            datas=data,
        )
        ir_attach.create(values)

    def cleanup(self):
        shutil.rmtree(self.output_dir)
        os.remove(self.output_dir + '.tar')

    def set_initial_queryset(self):
        pass

    def get_context(self):
        pass


class SBH(Creator):
    def set_initial_queryset(self):
        q = [('po_num', '=', self.po_num), ]

        self.qs_po = self.env['outsource.purchase.order'].search(q, limit=1)

        # queryset for all resources filtered by po_num
        # RPT01_ Monthly Accruals - Mobillized

        self.qs_invoice = self.env['outsource.invoice'].search(
            [('invoice_date', '=', '{}-{}-{}'.format(self.period_end.year, self.period_end.month, 1)),
             ('po_line_detail_id.po_line_id.po_id.po_num', '=', self.po_num)],
            order='invoice_date desc')

        resource_ids = self.qs_invoice.mapped('resource_id.id')
        self.qs_resource = self.env['outsource.resource'].search([('id', 'in', resource_ids)])
        if self.division:
            self.qs_resource = self.qs_resource.filtered(lambda r: r.division == self.division)

        # queryset for all PO Line filtered by po_num
        q = []
        q += [('po_line_id', 'in', self.qs_po.mapped('po_line_ids.id'))]

        self.qs_line_details = self.env['outsource.purchase.order.line.detail'].search(q)
        if self.division:
            self.qs_line_details = self.qs_line_details.filtered(lambda r: r.division == self.division)

        # get unit price
        q = [('contractor', '=', self.qs_po.contractor)]

        self.qs_unit_price = self.env['outsource.unit.rate'].search(q)

    def get_context(self):
        # obj is a single PurchaseOrder model

        df_resource = pd.DataFrame(
            odoo_to_pandas_list(self.qs_resource, ['id', 'po_os_ref', 'agency_ref_num', 'res_full_name',
                                                   'po_position', 'po_level', 'date_of_join',
                                                   'manager', 'director',
                                                   'po_line_detail_id.rate_diff_percent_calculated',
                                                   'has_tool_or_uniform']
                                )
        )

        df_resource['period_start'] = self.period_start
        df_resource['period_end'] = self.period_end
        df_resource['required_hour'] = self.required_hour
        df_resource['po_position'] = df_resource['po_position'].apply(lambda x: x.upper())
        df_invoice = pd.DataFrame(
            odoo_to_pandas_list(self.qs_invoice, ['id', 'resource_id.id', 'invoice_hour', 'invoice_claim',
                                                  'remarks']
                                )
        )

        if df_invoice.empty:
            df_invoice = pd.DataFrame(columns=['resource_id.id', 'id', 'invoice_claim', 'invoice_hour', 'remarks'])
        else:
            df_invoice = df_invoice.groupby('resource_id.id').first().reset_index()

        rs_resource = pd.merge(left=df_resource, right=df_invoice, how='left', left_on='id', right_on='resource_id.id')
        rs_resource.sort_values(
            by=['director', 'manager', 'po_position', 'po_line_detail_id.rate_diff_percent_calculated'],
            ascending=[True, True, True, True],
            inplace=True)
        rs_resource = rs_resource.fillna(0.0).to_dict('records')

        # to get rs_summary

        df_line_details = pd.DataFrame(
            odoo_to_pandas_list(self.qs_line_details,
                                ['po_position', 'po_os_ref', 'po_level', 'rate_diff_percent_calculated']))
        df_line_details['po_position'] = df_line_details['po_position'].apply(lambda x: x.upper())
        df_line_details = df_line_details.groupby(
            ['po_os_ref', 'po_position', 'po_level', 'rate_diff_percent_calculated']). \
            size().reset_index()
        df_line_details.columns = ['po_os_ref', 'po_position', 'po_level', 'rate_diff_percent_calculated', 'count']
        df_line_details = df_line_details.pivot_table(
            index=['po_position', 'po_os_ref', 'rate_diff_percent_calculated'],
            columns='po_level',
            values='count')
        rs_summary = df_line_details.reset_index()
        rs_summary.sort_values(by=['po_os_ref', 'po_position', 'rate_diff_percent_calculated'],
                               ascending=[True, True, True],
                               inplace=True)

        # make all columns lower
        rs_summary.columns = [i.lower() for i in rs_summary.columns]

        rs_summary = rs_summary.fillna(0.0).to_dict('records')
        df_unit_price = pd.DataFrame(odoo_to_pandas_list(self.qs_unit_price))
        df_unit_price = pd.pivot_table(df_unit_price, index=['po_position'], columns='po_level', values="amount"). \
            reset_index()
        rs_unit_price = df_unit_price.to_dict('records')

        context = {
            'contractor': self.qs_po.contractor,
            'po_num': self.qs_po.po_num,
            'po_line_count': self.qs_po.line,
            'rs_resource': rs_resource,  # recordset for individual resource (employee information)
            'rs_summary': rs_summary,  # recordset for summary of total numbers
            'rs_unit_price': rs_unit_price,  # recordset for summary of unit price
            'period_string': self.period_string,
            'total_required_hour': self.required_hour
        }

        return context

    def single_set_sbh(self, context=None):
        wb = load_workbook(self.form)

        # WORK SHEET SBH-FORM 1
        # ---------------------------------------------------------------------------------------
        row = 15
        column = 2
        sr = 1
        ws = wb.get_sheet_by_name('SBH-FORM 1')

        ws.cell('D5').value = context['contractor']
        ws.cell('G5').value = context['po_num']
        ws.cell('G7').value = 1
        ws.cell('D7').value = context['period_string']

        # CELLS TO BE UNLOCK
        unlock_cells = [
            'J39', 'L39',
            'J41', 'L41',
            'J43',

            'G46', 'J46', 'L46', 'M46',

            'C49', 'D49',

            'C53', 'C53',
            'C54', 'C54',
            'C55', 'C55',
            'C56', 'C56',
        ]

        rs_resource = context['rs_resource']

        ws.insert_rows(row, len(rs_resource) - 1)

        for record in rs_resource:
            ws.cell(row=row, column=column).value = sr
            ws.cell(row=row, column=column + 1).value = record['po_os_ref']
            ws.cell(row=row, column=column + 2).value = record['agency_ref_num']
            ws.cell(row=row, column=column + 3).value = record['res_full_name']
            ws.cell(row=row, column=column + 4).value = record['po_position']
            ws.cell(row=row, column=column + 5).value = record['po_level']
            ws.cell(row=row, column=column + 6).value = '' if not record['date_of_join'] else '{0:%d-%b-%Y}'.format(
                fields.Date.from_string(record['date_of_join']))
            ws.cell(row=row, column=column + 7).value = record['po_line_detail_id.rate_diff_percent_calculated'] / 100
            ws.cell(row=row, column=column + 8).value = record['required_hour']
            ws.cell(row=row, column=column + 9).value = record['invoice_claim']
            ws.cell(row=row, column=column + 10).value = '' if record['has_tool_or_uniform'] in ['false', 'FALSE',
                                                                                                 False,
                                                                                                 'False'] else 'Yes'
            ws.cell(row=row, column=column + 11).value = '' if len(str(record['remarks'])) == 0 else record['remarks']
            ws.cell(row=row, column=column + 12).value = record['manager']
            ws.cell(row=row, column=column + 13).value = record['director']
            ws.cell(row=row, column=column + 14).value = ''

            # # Unlock Cells
            ws.cell(row=row, column=column + 9).protection = Protection(locked=False)
            ws.cell(row=row, column=column + 11).protection = Protection(locked=False)
            ws.cell(row=row, column=column + 12).protection = Protection(locked=False)
            ws.cell(row=row, column=column + 13).protection = Protection(locked=False)
            ws.cell(row=row, column=column + 14).protection = Protection(locked=False)

            sr += 1
            row += 1
        # ---------------------------------------------------------------------------------------
        # END SBH FORM 1

        # WORK SHEET SBH-FORM 2
        # ---------------------------------------------------------------------------------------
        row = 16
        column = 2  # B

        ws = wb.get_sheet_by_name('SBH-FORM 2')

        ws.insert_rows(row, len(context['rs_summary']) - 1)
        ws.cell('R13').value = round(context['total_required_hour'])

        for record in context['rs_summary']:
            ws.cell(row=row, column=column).value = record['po_os_ref']
            ws.cell(row=row, column=column + 1).value = record.get('po_position')
            ws.cell(row=row, column=column + 2).value = record.get('level 1', 0)
            ws.cell(row=row, column=column + 3).value = record.get('level 2', 0)
            ws.cell(row=row, column=column + 4).value = record.get('level 3', 0)
            ws.cell(row=row, column=column + 12).value = record.get('rate_diff_percent_calculated', 0) / 100

            row += 1
        # ---------------------------------------------------------------------------------------
        # END SBH FORM 2

        # WORK SHEET UNIT PRICE
        # ---------------------------------------------------------------------------------------
        row = 1
        column = 1  # B

        ws = wb.get_sheet_by_name('UNIT PRICE')

        for record in context['rs_unit_price']:
            ws.cell(row=row, column=column).value = record['po_position']
            ws.cell(row=row, column=column + 1).value = record.get('Level 1', 0)
            ws.cell(row=row, column=column + 2).value = record.get('Level 2', 0)
            ws.cell(row=row, column=column + 3).value = record.get('Level 3', 0)
            ws.cell(row=row, column=column + 4).value = record.get('Level 4', 0)

            row += 1
        # ---------------------------------------------------------------------------------------
        # END UNIT PRICE

        # UNLOCK CELLS
        # ---------------------------------------------------------------------------------------
        for unlock_cell in unlock_cells:
            u_cell = ws.cell(unlock_cell)
            ws.cell(row=u_cell.row + sr, column=u_cell.col_idx).protection = Protection(locked=False)

        # ---------------------------------------------------------------------------------------
        # END UNLOCK CELLS

        for sheet in wb.worksheets:
            sheet.protection.enable()
            sheet.protection.set_password(self.xlsx_pass)
        wb.save(os.path.join(self.output_dir, self.output_filename))

    def make_sbh_per_po(self, po_num=None, period_start=None, period_end=None,
                        ramadan_start=None, ramadan_end=None, total_required_hours=None):
        self.po_num = po_num
        self.period_start = period_start
        self.period_end = period_end
        self.ramadan_start = ramadan_start
        self.ramadan_end = ramadan_end
        self.total_required_hours = total_required_hours
        self.setup()
        self.set_initial_queryset()

        try:
            context = self.get_context()
            self.output_filename = get_output_file(
                '{:0>2}'.format(self.period_start.month),
                context['contractor'],
                context['po_num'],
                context['period_string'],
            )
            self.single_set_sbh(context)
        except Exception as e:
            print('Failed: ', po_num, ' due to missing {}'.format(str(e)))

    def make_sbh_per_contractor(self, contractor=None, period_start=None, period_end=None,
                                ramadan_start=None, ramadan_end=None, total_required_hours=None):
        if contractor.lower() == 'all':
            qs = self.env['outsource.purchase.order'].search([])
        else:
            qs = self.env['outsource.purchase.order'].search([('contractor', '=', contractor)])

        for i in qs:
            self.make_sbh_per_po(i.po_num, period_start, period_end, ramadan_start, ramadan_end, total_required_hours)

    def make_sbh_per_division(self, po_num=None, division=None, period_start=None, period_end=None,
                              ramadan_start=None, ramadan_end=None, total_required_hours=None):
        self.division = division
        self.po_num = po_num
        self.period_start = period_start
        self.period_end = period_end
        self.ramadan_start = ramadan_start
        self.ramadan_end = ramadan_end
        self.total_required_hours = total_required_hours
        self.setup()
        self.set_initial_queryset()
        try:
            context = self.get_context()
            self.output_filename = get_output_file(
                '{:0>3}'.format(self.period_start.month),
                self.division or '',
                context['contractor'],
                context['po_num'],
                context['period_string'],
            )
            self.single_set_sbh(context)
        except Exception as e:
            print('Failed: ', po_num, ' due to missing {}'.format(str(e)))


class DATASHEET(Creator):
    def __init__(self, *args, **kwargs):
        super(DATASHEET, self).__init__(*args, **kwargs)
        self.contractor = None

    def set_initial_queryset(self):
        q = [('contractor', '=', self.contractor)]
        self.qs_resource = self.env['outsource.resource'].search(q)

    def get_context(self):
        df_resource = pd.DataFrame(odoo_to_pandas_list(self.qs_resource,
                                                       ['division', 'section', 'manager', 'director', 'access_db_id',
                                                        'po_line_detail_id.access_db_id',
                                                        'po_line_detail_id.po_rate',
                                                        'po_line_detail_id.rate',
                                                        'po_line_detail_id.rate_diff_percent_calculated',
                                                        'po_num', 'agency_ref_num', 'res_full_name',
                                                        'res_job_title', 'grade_level',
                                                        'date_of_join', 'has_tool_or_uniform']))

        df_resource.sort_values(
            by=['division', 'manager', 'director', 'res_job_title'],
            ascending=[True, True, True, True],
            inplace=True)
        rs_resource = df_resource.to_dict('records')

        context = {
            'contractor': self.contractor,
            'rs_resource': rs_resource,  # recordset for individual resource (employee information)
            'period_string': self.period_string,
        }

        return context

    def single_set_datasheet(self, context=None):

        wb = load_workbook(self.form)

        # WORK SHEET SBH-FORM 1
        # ---------------------------------------------------------------------------------------
        row = 7
        column = 2
        sr = 1
        ws = wb.get_sheet_by_name('STAFF LIST')

        ws.cell('B1').value = "713H CLAIMS FOR {}".format(context['contractor'])
        ws.cell('B2').value = context['period_string']

        rs_resource = context['rs_resource']

        ws.insert_rows(row, len(rs_resource) - 1)

        for record in rs_resource:
            ws.cell(row=row, column=column).value = sr
            ws.cell(row=row, column=column + 1).value = record.get('division', '')
            ws.cell(row=row, column=column + 2).value = record.get('section', '')
            ws.cell(row=row, column=column + 3).value = record.get('manager', '')
            ws.cell(row=row, column=column + 4).value = record.get('director', '')
            ws.cell(row=row, column=column + 5).value = record.get('access_db_id', '')
            ws.cell(row=row, column=column + 6).value = record.get('po_line_detail_id.access_db_id', '')
            ws.cell(row=row, column=column + 7).value = record.get('po_num', '')
            ws.cell(row=row, column=column + 8).value = record.get('po_line_detail_id.po_rate', '')
            ws.cell(row=row, column=column + 9).value = record.get('po_line_detail_id.rate_diff_percent_calculated', '')
            ws.cell(row=row, column=column + 10).value = record.get('po_line_detail_id.rate', '')
            ws.cell(row=row, column=column + 11).value = record.get('agency_ref_num', '')
            ws.cell(row=row, column=column + 12).value = record.get('res_full_name', '').title()
            ws.cell(row=row, column=column + 13).value = record.get('res_job_title', '')
            ws.cell(row=row, column=column + 14).value = record.get('grade_level', '')
            ws.cell(row=row, column=column + 15).value = '' if not record['date_of_join'] else '{0:%d-%b-%Y}'.format(
                fields.Date.from_string(record['date_of_join']))
            ws.cell(row=row, column=column + 16).value = 'Yes' if record.get('has_tool_or_uniform', '') else ""
            ws.cell(row=row, column=column + 17).value = self.required_hour
            ws.cell(row=row, column=column + 18).value = self.required_days

            # # Unlock Cells
            ws.cell(row=row, column=column + 20).protection = Protection(locked=False)
            ws.cell(row=row, column=column + 21).protection = Protection(locked=False)

            sr += 1
            row += 1
        # ---------------------------------------------------------------------------------------
        # END SBH FORM 1

        row_auto_filter = row - 1
        ws.auto_filter.ref = "B6:T{}".format(row_auto_filter)
        ws.protection.autoFilter = False

        ws1 = wb.get_sheet_by_name('MISMATCH')
        ws1.auto_filter.ref = "B6:R{}".format(row_auto_filter)
        ws1.protection.autoFilter = False
        ws1.protection.insertRows = False

        for sheet in wb.worksheets:
            sheet.protection.enable()
            sheet.protection.set_password(self.xlsx_pass)
        wb.save(os.path.join(self.output_dir, self.output_filename))

    def make_datasheet(self, contractor=None, period_start=None, period_end=None,
                       ramadan_start=None, ramadan_end=None, total_required_hours=None):
        self.contractor = contractor
        self.period_start = period_start
        self.period_end = period_end
        self.ramadan_start = ramadan_start
        self.ramadan_end = ramadan_end
        self.total_required_hours = total_required_hours
        self.setup()
        self.set_initial_queryset()
        try:
            context = self.get_context()
            self.output_filename = get_output_file(
                'DATASHEET',
                context['contractor'],
                context['period_string']
            )
            self.single_set_datasheet(context)
        except Exception as e:
            print('Failed: ', contractor, ' due to missing {}'.format(str(e)))

    def make_datasheet_per_contractor(self, contractor=None, period_start=None, period_end=None,
                                      ramadan_start=None, ramadan_end=None, total_required_hours=None):
        if contractor.lower() == 'all':
            contractors_tuple = get_distinct_selection(self, model='outsource.resource', field_name='contractor')
            contractors = [i[0] for i in contractors_tuple]
            for contractor in contractors:
                self.make_datasheet(contractor, period_start, period_end,
                                    ramadan_start, ramadan_end, total_required_hours)
        else:
            self.make_datasheet(contractor, period_start, period_end,
                                ramadan_start, ramadan_end, total_required_hours)
